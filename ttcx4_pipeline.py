#!/usr/bin/env python3
import argparse
import json
import re
import sys
import time
from collections import Counter
from itertools import product
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import chisquare

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
URL_TMPL = "https://caipiao.eastmoney.com/Result/History/ttcx4?page={page}"
HEADERS = ["page", "page_row", "期号", "开奖日期", "数字1", "数字2", "数字3", "数字4", "开奖号码"]
POSITIONS = ["第1位", "第2位", "第3位", "第4位"]
DIGITS = [str(i) for i in range(10)]


class PipelineError(Exception):
    pass


def fetch_page(page: int) -> str:
    url = URL_TMPL.format(page=page)
    last_err = None
    for attempt in range(4):
        try:
            req = Request(url, headers={"User-Agent": UA, "Referer": "https://caipiao.eastmoney.com/"})
            with urlopen(req, timeout=30) as resp:
                return resp.read().decode("utf-8", "ignore")
        except Exception as e:
            last_err = e
            time.sleep(0.8 * (attempt + 1))
    raise PipelineError(f"page={page} 获取失败: {last_err}")


def detect_total_pages() -> int:
    html = fetch_page(1)
    page_matches = [int(item) for item in re.findall(r"\?page=(\d+)", html)]
    if not page_matches:
        raise PipelineError("未能从第1页分页导航中识别总页数，请检查网页结构是否变化")
    total_pages = max(page_matches)
    if total_pages < 1:
        raise PipelineError(f"识别到的总页数异常: {total_pages}")
    return total_pages


def parse_page(html: str, page: int):
    tbody_match = re.search(r"<tbody>(.*?)</tbody>", html, re.S)
    if not tbody_match:
        raise PipelineError(f"page={page} 未找到 tbody")
    row_html_list = re.findall(r"<tr>(.*?)</tr>", tbody_match.group(1), re.S)
    if not row_html_list:
        raise PipelineError(f"page={page} 未解析到开奖记录")
    rows = []
    for row_idx, row_html in enumerate(row_html_list, 1):
        cells = re.findall(r"<td.*?>(.*?)</td>", row_html, re.S)
        if len(cells) < 4:
            raise PipelineError(f"page={page} row={row_idx} 单元格数量异常")
        issue_m = re.search(r">(\d+)<", cells[0])
        date = re.sub(r"<.*?>", "", cells[1]).strip()
        nums = re.findall(r"<span class=\"pellet[^>]*?red\">(\d+)</span>", cells[3])
        if not issue_m or len(nums) != 4:
            raise PipelineError(f"page={page} row={row_idx} 开奖号码解析异常")
        issue = issue_m.group(1)
        rows.append([page, row_idx, issue, date, nums[0], nums[1], nums[2], nums[3], "".join(nums)])
    return rows


def fetch_all(page_count: int):
    all_rows = []
    page_counts = []
    for page in range(1, page_count + 1):
        html = fetch_page(page)
        rows = parse_page(html, page)
        all_rows.extend(rows)
        page_counts.append((page, len(rows)))
        time.sleep(0.05)
    return all_rows, page_counts


def compute_sheet_ranges(total_pages: int, chunk: int = 40):
    if total_pages < 1:
        raise PipelineError(f"总页数异常: {total_pages}")
    ranges = []
    start = 1
    while start <= total_pages:
        end = min(start + chunk - 1, total_pages)
        ranges.append((f"{start}-{end}页", start, end))
        start = end + 1
    return ranges


def validate_fetch(all_rows, page_counts, expected_pages: int):
    page_to_count = {}
    for page, count in page_counts:
        if page in page_to_count:
            raise PipelineError(f"页数校验失败：page={page} 重复出现")
        page_to_count[page] = count

    missing_pages = [page for page in range(1, expected_pages + 1) if page not in page_to_count]
    if missing_pages:
        preview = ", ".join(map(str, missing_pages[:10]))
        suffix = " ..." if len(missing_pages) > 10 else ""
        raise PipelineError(f"页数校验失败：缺少页面 {preview}{suffix}")

    empty_pages = [page for page, count in page_counts if count <= 0]
    if empty_pages:
        raise PipelineError(f"页数校验失败：以下页面没有解析到记录 {empty_pages}")

    issue_counter = Counter(row[2] for row in all_rows)
    duplicate_issues = sorted(issue for issue, count in issue_counter.items() if count > 1)
    if duplicate_issues:
        preview = ", ".join(duplicate_issues[:10])
        suffix = " ..." if len(duplicate_issues) > 10 else ""
        raise PipelineError(f"期号校验失败：发现重复期号 {preview}{suffix}")

    baseline_samples = [count for _, count in page_counts[:-1]] or [page_counts[0][1]]
    baseline_count = Counter(baseline_samples).most_common(1)[0][0]
    abnormal_pages = [(page, count) for page, count in page_counts[:-1] if count != baseline_count]
    if abnormal_pages:
        preview = ", ".join(f"page={page}:{count}条" for page, count in abnormal_pages[:10])
        suffix = " ..." if len(abnormal_pages) > 10 else ""
        raise PipelineError(f"页内记录数校验失败：非末页记录数异常 {preview}{suffix}")

    warnings = []
    page_status = {}
    for page, count in page_counts:
        if page == expected_pages and count != baseline_count:
            page_status[page] = f"末页 {count} 条"
            warnings.append(f"末页记录数为 {count}，常规页为 {baseline_count}，按历史末页处理")
        else:
            page_status[page] = "正常"

    return {
        "page_status": page_status,
        "warnings": warnings,
        "baseline_count": baseline_count,
    }


def apply_basic_style(ws, center=True):
    base_font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = base_font
            cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    for cell in ws[1]:
        cell.font = header_font


def write_main_excel(all_rows, page_counts, page_status, target_file: Path, total_pages: int, fetch_warnings):
    wb = Workbook()
    wb.remove(wb.active)
    widths = {1: 8, 2: 10, 3: 12, 4: 20, 5: 8, 6: 8, 7: 8, 8: 8, 9: 12}
    sheet_ranges = compute_sheet_ranges(total_pages)

    for title, start, end in sheet_ranges:
        ws = wb.create_sheet(title)
        ws.append(HEADERS)
        for row in all_rows:
            if start <= row[0] <= end:
                ws.append(row)
        apply_basic_style(ws, center=True)
        for idx, width in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = "A2"

    ws = wb.create_sheet("页数校验")
    ws.append(["page", "记录数", "状态"])
    for page, count in page_counts:
        ws.append([page, count, page_status.get(page, "正常")])
    apply_basic_style(ws, center=True)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.freeze_panes = "A2"

    latest = all_rows[0]
    ws = wb.create_sheet("说明")
    ws.append(["说明"])
    ws.append([f"已更新至最新一期：{latest[2]}，日期 {latest[3]}，开奖号码 {latest[8]}。"])
    ws.append([f"当前按东方财富历史页共 {total_pages} 页重新全量抓取。"])
    ws.append([f"为避免某些查看器对单个大工作表的预览截断，数据按页码动态拆分为：{'、'.join(title for title, _, _ in sheet_ranges)}。"])
    if fetch_warnings:
        for warning in fetch_warnings:
            ws.append([f"抓取校验提示：{warning}"])
    else:
        ws.append(["抓取完整性校验通过：页数齐全、非末页记录数一致、无重复期号。"])
    apply_basic_style(ws, center=False)
    ws.column_dimensions["A"].width = 120

    wb.save(target_file)


def analyze_numbers(nums_only):
    n = len(nums_only)
    expected = n / 10
    counts = {p: Counter() for p in POSITIONS}
    for num in nums_only:
        for i, p in enumerate(POSITIONS):
            counts[p][num[i]] += 1

    chi_results = []
    for p in POSITIONS:
        obs = [counts[p][d] for d in DIGITS]
        chi2, pv = chisquare(obs)
        chi_results.append({"position": p, "chi2": float(chi2), "pvalue": float(pv)})

    if all(item["pvalue"] >= 0.05 for item in chi_results):
        conclusion = "基本均匀，无明显偏差"
    elif any(item["pvalue"] < 0.01 for item in chi_results):
        conclusion = "某位置存在统计上显著偏差"
    else:
        conclusion = "某些位置存在轻微但不稳定偏差"

    position_notes = []
    for p in POSITIONS:
        arr = []
        for d in DIGITS:
            c = counts[p][d]
            diff = c - expected
            rate = diff / expected
            arr.append((abs(diff), diff, d, c, rate))
        arr.sort(reverse=True)
        top = []
        for _, diff, d, _, rate in arr[:2]:
            direction = "偏多" if diff > 0 else "偏少"
            top.append(f"{d}({direction} {diff:+.1f} 次, {rate:+.2%})")
        position_notes.append({"position": p, "note": "；".join(top)})

    return counts, expected, chi_results, conclusion, position_notes


def generate_top10(nums_only):
    n = len(nums_only)
    overall = [Counter(num[i] for num in nums_only) for i in range(4)]
    windows = [100, 300, 500]
    all_candidates = ["".join(t) for t in product("0123456789", repeat=4)]
    recent_sets = {w: set(nums_only[:w]) for w in windows}
    common_allowed = [num for num in all_candidates if all(num not in recent_sets[w] for w in windows)]

    score_maps = {}
    rank_maps = {}
    for w in windows:
        recent = nums_only[:w]
        recent_c = [Counter(num[i] for num in recent) for i in range(4)]
        pos_scores = []
        for i in range(4):
            m = {}
            for d in DIGITS:
                exp = w * (overall[i][d] / n)
                m[d] = exp - recent_c[i][d]
            pos_scores.append(m)
        score_maps[w] = pos_scores
        scored = []
        for num in common_allowed:
            score = sum(pos_scores[i][num[i]] for i in range(4))
            scored.append((score, num))
        scored.sort(reverse=True)
        rank_maps[w] = {num: idx + 1 for idx, (_, num) in enumerate(scored)}

    consensus = []
    for num in common_allowed:
        ranks = [rank_maps[w][num] for w in windows]
        scores = [sum(score_maps[w][i][num[i]] for i in range(4)) for w in windows]
        consensus.append(
            {
                "number": num,
                "top1000_windows": sum(r <= 1000 for r in ranks),
                "avg_rank": sum(ranks) / len(ranks),
                "min_score": min(scores),
                "sum_score": sum(scores),
                "r100": ranks[0],
                "r300": ranks[1],
                "r500": ranks[2],
                "s100": scores[0],
                "s300": scores[1],
                "s500": scores[2],
            }
        )
    consensus.sort(key=lambda x: (-x["top1000_windows"], x["avg_rank"], -x["min_score"], -x["sum_score"], x["number"]))

    selected = []
    used_first2 = set()
    used_last2 = set()
    for item in consensus:
        num = item["number"]
        if num[:2] in used_first2 or num[2:] in used_last2:
            continue
        if any(sum(a != b for a, b in zip(num, old["number"])) < 3 for old in selected):
            continue
        selected.append(item)
        used_first2.add(num[:2])
        used_last2.add(num[2:])
        if len(selected) == 10:
            break

    if len(selected) < 10:
        for item in consensus:
            num = item["number"]
            if num in [old["number"] for old in selected]:
                continue
            if any(sum(a != b for a, b in zip(num, old["number"])) < 2 for old in selected):
                continue
            selected.append(item)
            if len(selected) == 10:
                break

    for idx, item in enumerate(selected, 1):
        item["rank"] = idx
        item["note"] = "跨100/300/500期窗口共识 + 分散约束"
    return selected


def write_analysis_excel(all_rows, analysis_file: Path):
    nums_only = [row[8] for row in all_rows]
    counts, expected, chi_results, conclusion, position_notes = analyze_numbers(nums_only)
    top10 = generate_top10(nums_only)

    base_font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True)
    red_fill = PatternFill("solid", fgColor="FDE9D9")
    green_fill = PatternFill("solid", fgColor="E2F0D9")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")

    wb = Workbook()

    ws = wb.active
    ws.title = "汇总结论"
    ws.append(["项目", "结果"])
    summary_rows = [
        ("样本期数", len(nums_only)),
        ("最新一期", all_rows[0][2]),
        ("最新开奖日期", all_rows[0][3]),
        ("最新开奖号码", all_rows[0][8]),
        ("每个位置理论期望次数", expected),
        ("总体结论", conclusion),
        ("说明", "卡方检验原假设：该位置上0-9均匀分布"),
    ]
    for row in summary_rows:
        ws.append(list(row))
    for item in position_notes:
        ws.append([item["position"] + "最明显偏差", item["note"]])
    for row in ws.iter_rows():
        for cell in row:
            cell.font = base_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 90

    ws = wb.create_sheet("位置数字频数")
    ws.append(["位置"] + DIGITS + ["合计"])
    for p in POSITIONS:
        ws.append([p] + [counts[p][d] for d in DIGITS] + [sum(counts[p][d] for d in DIGITS)])
    apply_basic_style(ws, center=True)
    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 10
    ws.freeze_panes = "A2"

    ws = wb.create_sheet("偏离率")
    ws.append(["位置"] + DIGITS)
    for p in POSITIONS:
        ws.append([p] + [(counts[p][d] - expected) / expected for d in DIGITS])
    apply_basic_style(ws, center=True)
    for r in range(2, 6):
        for c in range(2, 12):
            cell = ws.cell(r, c)
            cell.number_format = "0.00%"
            if cell.value >= 0.03:
                cell.fill = red_fill
            elif cell.value <= -0.03:
                cell.fill = green_fill
    for c in range(1, 12):
        ws.column_dimensions[get_column_letter(c)].width = 10
    ws.freeze_panes = "A2"

    ws = wb.create_sheet("卡方检验")
    ws.append(["位置", "卡方统计量", "p值", "判断"])
    for item in chi_results:
        pv = item["pvalue"]
        if pv < 0.01:
            judge = "显著偏差"
        elif pv < 0.05:
            judge = "轻微偏差"
        else:
            judge = "无明显偏差"
        ws.append([item["position"], item["chi2"], pv, judge])
    apply_basic_style(ws, center=True)
    for r in range(2, 6):
        ws.cell(r, 2).number_format = "0.000"
        ws.cell(r, 3).number_format = "0.000000"
        if ws.cell(r, 3).value < 0.01:
            for c in range(1, 5):
                ws.cell(r, c).fill = red_fill
        elif ws.cell(r, 3).value < 0.05:
            for c in range(1, 5):
                ws.cell(r, c).fill = yellow_fill
    for col, width in {"A": 10, "B": 14, "C": 12, "D": 14}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    ws = wb.create_sheet("偏多偏少")
    ws.append(["位置", "数字", "实际次数", "理论次数", "差值", "偏离率", "方向"])
    for p in POSITIONS:
        items = []
        for d in DIGITS:
            count = counts[p][d]
            diff = count - expected
            rate = diff / expected
            direction = "偏多" if diff > 0 else ("偏少" if diff < 0 else "持平")
            items.append((abs(diff), d, count, diff, rate, direction))
        items.sort(reverse=True)
        for _, d, count, diff, rate, direction in items:
            ws.append([p, d, count, expected, diff, rate, direction])
    apply_basic_style(ws, center=True)
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 4).number_format = "0.0"
        ws.cell(r, 5).number_format = "0.0"
        ws.cell(r, 6).number_format = "0.00%"
        if ws.cell(r, 6).value >= 0.03:
            for c in range(1, 8):
                ws.cell(r, c).fill = red_fill
        elif ws.cell(r, 6).value <= -0.03:
            for c in range(1, 8):
                ws.cell(r, c).fill = green_fill
    for col, width in {"A": 10, "B": 8, "C": 12, "D": 12, "E": 10, "F": 10, "G": 10}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    ws = wb.create_sheet("明日10组")
    ws.append(["排名", "号码", "r100", "r300", "r500", "平均排名", "top1000窗口数", "s100", "s300", "s500", "说明"])
    for item in top10:
        ws.append([
            item["rank"], item["number"], item["r100"], item["r300"], item["r500"], item["avg_rank"],
            item["top1000_windows"], item["s100"], item["s300"], item["s500"], item["note"]
        ])
    apply_basic_style(ws, center=True)
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 6).number_format = "0.0"
        ws.cell(r, 8).number_format = "0.00"
        ws.cell(r, 9).number_format = "0.00"
        ws.cell(r, 10).number_format = "0.00"
    for col, width in {"A": 8, "B": 10, "C": 8, "D": 8, "E": 8, "F": 12, "G": 14, "H": 10, "I": 10, "J": 10, "K": 34}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    wb.save(analysis_file)
    return {
        "sample_count": len(nums_only),
        "expected_per_digit": expected,
        "conclusion": conclusion,
        "chi_square": chi_results,
        "top10": top10,
    }


def run(data_dir: Path, pages: int = None):
    data_dir.mkdir(parents=True, exist_ok=True)
    total_pages = pages if pages is not None else detect_total_pages()
    if total_pages < 1:
        raise PipelineError(f"抓取页数异常: {total_pages}")

    main_file = data_dir / "ttcx4_历史开奖号码.xlsx"
    analysis_file = data_dir / "ttcx4_位置数字偏差分析.xlsx"

    all_rows, page_counts = fetch_all(total_pages)
    if not all_rows:
        raise PipelineError("未抓取到任何开奖记录")

    validation = validate_fetch(all_rows, page_counts, total_pages)
    latest = all_rows[0]
    write_main_excel(
        all_rows,
        page_counts,
        validation["page_status"],
        main_file,
        total_pages,
        validation["warnings"],
    )
    analysis = write_analysis_excel(all_rows, analysis_file)

    result = {
        "total_pages": total_pages,
        "latest_period": latest[2],
        "latest_date": latest[3],
        "latest_number": latest[8],
        "total_records": len(all_rows),
        "main_excel": str(main_file),
        "analysis_excel": str(analysis_file),
        "conclusion": analysis["conclusion"],
        "chi_square_summary": analysis["chi_square"],
        "fetch_warnings": validation["warnings"],
        "top10": analysis["top10"],
    }
    return result


def parse_args():
    parser = argparse.ArgumentParser(description="TTCX4 lottery fetch/update/analyze pipeline")
    parser.add_argument("--data-dir", default=".", help="数据目录，默认当前目录")
    parser.add_argument("--pages", type=int, default=None, help="抓取页数，默认自动识别总页数")
    return parser.parse_args()


def main():
    args = parse_args()
    try:
        result = run(Path(args.data_dir).expanduser().resolve(), args.pages)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False))
        sys.exit(1)


if __name__ == "__main__":
    main()
